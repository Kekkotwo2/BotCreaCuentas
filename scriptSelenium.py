from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
import time
from selenium.webdriver.common.action_chains import ActionChains
from openpyxl import load_workbook, Workbook
from datetime import datetime
import random

def create_email(driver,correo):
    #Abrimos una nueva pestaña con la nueva pag
    try:
        driver.execute_script("window.open('https://yopmail.com/es/', '_blank');")
    except Exception as e:
        print("Ha ocurrido un error al obtener la url")
        print(f"Error: {e}")
        return False
    time.sleep(2)

    #Vamos a la pestaña previamente abierta
    driver.switch_to.window(driver.window_handles[1])

    #Se ingresa el nombre del correo el cual pusimos en registrar
    try:
        WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, "//input[@id='login']"))
        ).clear()

        WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, "//input[@id='login']"))
        ).send_keys(correo)
    except Exception as e:
        print("Ha ocurrido un error en la creacion del correo temporal")
        print(f"Error: {e}")
        return False

    try:
        #Refrescamos para que se actualize la lista de correos
        WebDriverWait(driver, 20).until(
        EC.element_to_be_clickable((By.XPATH, "//*[@id='refreshbut']/button"))
        ).click()
        print("a")
        #Ingresamos al iframe
        wait = WebDriverWait(driver, 20)
        iframe = wait.until(EC.presence_of_element_located((By.ID, "ifmail")))
        driver.switch_to.frame(iframe)
        print("b")

        #Obtenemos el codigo que estaba en el body del correo
        print(correo)
        clave = wait.until(EC.visibility_of_element_located((By.XPATH, "//div[@id='mail']//b")))
        print("Creacion del Email Exitosa")
        return clave.text
    except Exception as e:
        print("Error al obtener el codigo de registro\nFuncion create_mail")
        print(f"Error: {e}")
        return False

def register(driver,correo,extension):
    
    try:
        #Es la pagina a la que entraremos
        driver.get("https://platforms.mapup.ai/login?flow=signup&firstName=API&lastName=1")
    except Exception as e:
        print("Ha ocurrido un error obteniendo la url")
        print(f"Error: {e}")
        return False

    try:

        WebDriverWait(driver, 5).until(
        EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'Logout')]"))
        )
        visible = True
    except:
        visible = False

    if visible == True:
        try:
            #Para la pantalla en donde le pide deslogearse de nuevo
            WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'Logout')]"))
            ).click()

            WebDriverWait(driver, 10).until(
                EC.visibility_of_element_located((By.ID, "radix-:rf:"))
            )

            WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//button[contains(@data-sentry-element, 'Button') and contains(., 'Logout') and contains(@class, 'bg-secondary')]"))
            ).click()

            driver.get("https://platforms.mapup.ai/login?flow=signup&firstName=API&lastName=1")
        except Exception as e:
            print(f"Error: {e}")
            return False
    try:
        #Campos que solicita la pagina el cual es Nombre y Apellido
        # WebDriverWait(driver, 10).until(
        # EC.element_to_be_clickable((By.ID, "firstName"))
        # ).send_keys("API")

        # WebDriverWait(driver, 10).until(
        # EC.element_to_be_clickable((By.ID, "lastName"))
        # ).send_keys("1")

        # #Click boton para continuar
        # WebDriverWait(driver, 10).until(
        # EC.element_to_be_clickable((By.XPATH, "//*[@id='radix-:ra:-content-signup']/div/div/div/button"))
        # ).click()

        #Otros campos solicitados en el registro Email y contra
        WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.ID, "Email"))
        ).send_keys(f"{correo}{extension}")

        WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.ID, "password"))
        ).send_keys("@Cial2024")

        #Click boton para continuar
        WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, "//*[@id='radix-:ra:-content-signup']/div/div/div/div/button[2]"))
        ).click()
    except Exception as e:
        print("A ocurrido un error en el formulario de registro")
        print(f"Error: {e}")
        #return False

    #Traemos la clave que fue enviada al correo
    clave = create_email(driver,correo)
    print(f"clave obtenida: {clave}")

    #Cerramos la ventana
    driver.close()
    driver.switch_to.window(driver.window_handles[0])

    try:
        #Ingresamos la clave que nos fue enviada en el 2fa del registro
        WebDriverWait(driver, 150).until(
            EC.element_to_be_clickable((By.XPATH, "//*[@id='code']"))
        ).send_keys(clave)
        print("Registro Exitoso")
        return True
    except Exception as e:
        print("Ha ocurrido un error en el ingreso de la clave que se nos envio al correo")
        print(f"Error: {e}")
        return False

def login(driver,correo,extension):
    #Vamos a la pagina login
    try:
        driver.get("https://platforms.mapup.ai/login?flow=login&appState=eyJmbG93Ijoic3NvIiwicmVkaXJlY3RfdXJpIjoiaHR0cHM6Ly90b2xsZ3VydS5jb20vYXV0aGVudGljYXRlIiwiYXBwbGljYXRpb24iOiJ0b2xsZ3VydV93ZWIifQ==")
    except Exception as e:
        print("Ha ocurrido un error al obtener la pagina login")
        print(f"Error: {e}")
        return False
   
    try:
        #Ingresamos el correo y password en el login
        #aveces no hace click x eso .click()
        WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, "email"))
        ).click()

        WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, "email"))
        ).send_keys(f"{correo}{extension}")

        WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, "password"))
        ).send_keys("@Cial2024")

        WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, "//*[@id='radix-:ra:-content-login']/div/div[1]/div/button"))
        ).click()
    except Exception as e:
        print("Ha ocurrido un error en el proceso de login")
        print(f"Error: {e}")
        return False

    #Esto es por si pedia el 2fa al logearse, aun no lo puedo probar
    try:
        #Se intenta presionar el boton elegir perfil
        WebDriverWait(driver, 15).until(
        EC.element_to_be_clickable((By.XPATH, "//div[contains(@class, 'cursor-pointer') and contains(@class, 'animate-slide-in-bottom')]"))
        ).click()
    except:
        #Si no se va a createmail y se ingresa nuevamente al correo para recoger la nueva clave
        print("2fa")
        clave = create_email(driver,correo)

        try:
            WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//*[@id='code']"))
            ).send_keys(clave)
        except Exception as e:
            print("2fa fallo")
            print(f"Error: {e}")
            return False
        WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, "//div[contains(@class, 'cursor-pointer') and contains(@class, 'animate-slide-in-bottom')]"))
        ).click()

    print("Login Exitoso")
    return True
def createProfile(driver):
    #Los botones en esta parte de la pagina funcionan raro, aveces se le da click y no quedan precionados
    contador = 0
    while  True:
        try:
            individual_button = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, "//button[@aria-label='Toll API']"))
            )

            # Desplaza la vista hacia el botón
            driver.execute_script("arguments[0].scrollIntoView(true);", individual_button)

            # Simula el clic en el botón "API"
            actions = ActionChains(driver)
            actions.move_to_element(individual_button).click().perform()



            WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//button[contains(@class, 'bg-gradient-to-t') and contains(., 'Next')]"))
            ).click()
        except Exception as e:
            print("Error al presionar el boton api")
            print(f"Error: {e}")
            return False

        try:
            #Si esta este boton significa que paso de pestaña a la siguiente por ende se sale del bucle
            driver.find_element(By.CSS_SELECTOR, 'button[aria-label="Individual"]').click()
            break
        except:
            #Si no este se queda dentro del buqlue
            contador =+ 1

        if contador == 5:
            print("Se termino la ejecucion ya que se han hecho 5 intentos en boton api")
            return

    contador = 0
    while  True:
        try:
            # Espera hasta que el botón "Individual" esté presente y sea clickeable
            individual_button = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, "//button[@aria-label='Individual']"))
            )

            # Desplaza la vista hacia el botón
            driver.execute_script("arguments[0].scrollIntoView(true);", individual_button)

            # Simula el clic en el botón "Individual"
            actions = ActionChains(driver)
            actions.move_to_element(individual_button).click().perform()

            # Ahora busca el botón "Next"
            WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//button[contains(@class, 'bg-gradient-to-t') and contains(., 'Next')]"))
            ).click()

        except Exception as e:
            print("Error al presionar Boton enterprise")
            print(f"Error: {e}")
            return False

        #Si esta este input significa que paso de pestaña a la siguiente por ende se sale del bucle
        try:
            WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'input[name="phone_number"]'))
            )
            break

        except:
            contador =+ 1

        if contador == 5:
            print("Se termino la ejecucion ya que se han hecho 5 intentos en boton personal")
            return

    try:
        phone_input = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.NAME, 'phone_number')))

        driver.execute_script("arguments[0].scrollIntoView(true);", phone_input)

        lastNumbers = random.randint(1000000, 9999999)
        phone_input.clear()  # Limpia el campo si tiene algún valor
        phone_input.send_keys(f'+1555{lastNumbers}')  # Ingresa el número de teléfono

    except Exception as e:
        print("Ha ocurrido un error en el ingreso del numero de telefono")
        print(f"Error: {e}")
        return False

    #Inputs nombre y descripcion
    try:
        WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="linkedin_profile"]'))
        ).send_keys("cprofile")

        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="usage_intent"]'))
        ).send_keys("qwertyuiopqwertyuiopqwertyuiopqwertyuiopqwertyuiopqwertyuiopqwertyuiopqwertyuiopqwertyuiopqwertyuiopqwertyuiopqwertyuiop")


        WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, "//button[contains(@class, 'bg-gradient-to-t') and contains(., 'Next')]"))
        ).click()
    except Exception as e:
        print("Error en el ingreso de perfil")
        print(f"Error: {e}")
        return False

    try:
        # Espera hasta que el botón esté presente y sea clickeable
        start_trial_button = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.XPATH, "//button[contains(@class, 'bg-gradient-to-t') and contains(text(), 'Start Trial')]"))
        )

        # Desplazar hacia el botón
        driver.execute_script("arguments[0].scrollIntoView();", start_trial_button)

        # Hacer clic en el botón usando JavaScript
        driver.execute_script("arguments[0].click();", start_trial_button)

    except Exception as e:
        print("Error en start trial")
        print(f"Error: {e}")
        return False

    try:
       # Espera hasta que el modal sea visible
        WebDriverWait(driver, 20).until(
            EC.visibility_of_element_located((By.ID, "modal_header"))
        )

        # Espera hasta que el botón "Start Trial" dentro del modal esté clickeable
        start_trial_button = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, "//div[@id='modal_body']//button[contains(text(), 'Start Trial')]"))
        )

        start_trial_button.click()

    except Exception as e:
        print("Error en start trial dentro de modal")
        print(f"Error: {e}")
        return False
    print("Creacion de perfil Exitoso")
    return True
def getApi(driver):
    time.sleep(10)

    driver.execute_script("window.scrollTo(0, 0);")
    try:
        subscription_button = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, "//button[contains(@class, 'items-center') and contains(., 'Subscriptions')]"))
        )
        subscription_button.click()
    except Exception as e:
        print("Error al presionar boton subcriptions")
        print(f"Error: {e}")
        return False
    try:
        WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, "//a[contains(text(), 'API key')]"))
        ).click()
    except Exception as e:
        print("Error al presionar boton API key")
        print(f"Error: {e}")
        return False
    try:
        input_field = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.ID, "hero-field"))
        )
        api = input_field.get_attribute("value")
        print("El valor del campo es:", api)
        print("Obtencion de api Exitoso")
        return api
    except Exception as e:
        print("Se cayó el obtener API")
        print(f"Error: {e}")
        return None
def logOut(driver):
    try:
        # Esperar hasta que el botón sea clickeable
        button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, "headlessui-menu-button-:r9R6:"))
        )
        button.click()  # Hacer clic en el botón
    except Exception as e:
        print("Error al hacer clic en el botón:", e)

    try:
        # Esperar hasta que el botón de "Sign out" sea clickeable
        sign_out_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, "logout-button"))
        )
        sign_out_button.click()  # Hacer clic en el botón "Sign out"
    except Exception as e:
        print("Error al hacer clic en el botón 'Sign out':", e)
    return True
def overwriteExcelRows(lista,flag):
    #Cargamos el excel
    if flag == True:
        wb = load_workbook(filename = 'excelPruebaa.xlsx')
    else:
        wb = Workbook()
    ws = wb.active

    ws["a1"] = "Correo"
    ws["b1"] = "Api"
    ws["c1"] = "Fecha Inicio"


    #++++++++++++++++++++++++++++++++++++++++++++++++++++++
    #Esto es para insertar multiples filas por ejemplo si se requiere hacer ejecuciones de 10 hacer append a filas para luego recorrerlas y insertarlas
    #lista = [["Correo1","Api1",fecha],["Correo2","Api2",fecha]]

    for fila in lista:
        print(fila)
        ws.append(fila)
    #++++++++++++++++++++++++++++++++++++++++++++++++++++++
    wb.save(filename='excelPruebaa.xlsx')
    return True


def appendExcelUnique(correo, api):
    #Cargamos el excel
    wb = load_workbook(filename = 'excelPruebaa.xlsx')
    ws = wb.active


    now = datetime.now()
    fecha = now.strftime("%Y-%m-%d")

    #Creamos la fila para insertar, esta ejecucion es para registros unicos
    fila = [correo,api,fecha]

    ws.append(fila)
    #Guardamos el excel
    wb.save(filename='excelPruebaa.xlsx')
    print("Append exitoso")
    return True


def singleRowExecute():
    chrome_options = Options()
    # chrome_options.add_argument("--headless") # con esto lo ejecua sin que se abra el navegador
    chrome_options.add_argument("--incognito")
    chrome_options.add_argument("--start-maximized")
    contador = 0
    #correo = input("Ingrese un correo: ")
    lista = []
    now = datetime.now()
    fechaCompleta = now.strftime("%Y-%m-%d")

    añoActual = now.year
    mesActual = now.month
    diaActual = now.day
    extension = "@skynet.infos.st"
    for i in range(1):
        driver = webdriver.Chrome(options=chrome_options)

        #correo="correorealapi2"

        correo = f"correofalso45451{i}{diaActual}{mesActual}{añoActual}" #Aqui colocar nombre de correo

        #Si salta el error recuerde cambiar el correo
        if not register(driver,correo,extension):
            print("Proceso terminado con error")
            #sys.exit(1)
            break

        if not login(driver,correo,extension):
            print("Proceso terminado con error")
            #sys.exit(1)
            break

        if not createProfile(driver):
            print("Proceso terminado con error")
            #sys.exit(1)
            break

        api = getApi(driver)
        if api is None:
            print("Proceso terminado con error")
            #sys.exit(1)
            break

        lista.append([f"{correo}{extension}",api,fechaCompleta]) #el correo se va baneando
        # if not logOut(driver):
        #     print("Proceso terminado con error")
        #     sys.exit(1)

        driver.quit()
        contador += 1
        print(f"Iteracion n°: {contador}")
    overwriteExcelRows(lista,True)
    print("Fin")

if __name__ == "__main__":
    singleRowExecute()



#Agregar trys en las funciones, webdriverelement ✅✅
#Importar las claves a un excel acompañadas de un correo y fecha de creacion ✅✅
#probar varias ejecuciones +10 trayendo los correos desde un excel ✅
#que avise cuando ya existe
#Manejar el error del captcha
#aveces te pide un 2fa cuando logeas ✅
#si manda captcha en el yopmail se puede hjacer que vaya a otro correo temporal y repita el register